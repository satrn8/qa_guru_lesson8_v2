import zipfile
from PyPDF2 import PdfReader
from openpyxl import load_workbook
import csv
import os


def test_zip_file():
    zip_ = zipfile.ZipFile('.\\resources\\archive.zip', 'w')
    for folder, subfolders, files in os.walk('.\\folder\\'):
        for file in files:
            if file.endswith('.pdf') or file.endswith('.csv') or file.endswith('.xlsx'):
                zip_.write(os.path.join(folder, file),
                                  os.path.relpath(os.path.join(folder, file), '.\\resources\\'),
                                  compress_type=zipfile.ZIP_DEFLATED)

    zip_.close()


def test_unzip_files():
    zip_ = zipfile.ZipFile('.\\resources\\archive.zip')
    zip_.extractall('.\\unzip\\')
    zip_.close()


def test_read_pdf_file():
    reader = PdfReader(".\\folder\\pytest.pdf")
    page = reader.pages[0]
    text = page.extractText()
    assert "2022" in text


def test_read_xlsx_file():
    workbook = load_workbook(".\\folder\\file_example_XLSX_10.xlsx")
    sheet = workbook.active
    country = sheet.cell(row=4, column=5).value
    assert "France" == country


def test_read_csv_file():
    with open(".\\folder\\username.csv") as f:
        reader = csv.reader(f)
        headers = next(reader)
    assert "First name" in str(headers)


if __name__ == '__main__':
    test_zip_file()
    test_unzip_files()
    test_read_pdf_file()
    test_read_xlsx_file()
    test_read_csv_file()