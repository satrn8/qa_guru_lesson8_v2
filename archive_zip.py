from zipfile import ZipFile
from PyPDF2 import PdfReader
from openpyxl import load_workbook
import csv


def test_read_pdf_file():
    reader = PdfReader(".\\folder\\pytest.pdf")
    number_of_pages = len(reader.pages)
    print(number_of_pages)
    page = reader.pages[0]
    text = page.extractText()
    print(text)
    assert "2022" in text


def test_read_xlsx_file():
    workbook = load_workbook(".\\folder\\file_example_XLSX_10.xlsx")
    sheet = workbook.active
    country = sheet.cell(row=4, column=5).value
    assert "France" == country


def test_read_csv_file():
    with open(".\\folder\\username.csv") as f:
        reader = csv.reader(f)
        for i in reader:
            print(i)


def test_zip_file():
    pass